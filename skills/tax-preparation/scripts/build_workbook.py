#!/usr/bin/env python3
"""Framework for building auditable tax computation workbooks.

Creates an xlsx with:
- Source Data sheet: raw numbers from tax documents
- Tax Tables sheet: brackets, rates, thresholds from government sources
- Computation sheets: one per area (capital gains, rental, FTC, etc.)
- Federal/State Return sheets: one row per form line
- Form Values sheet: rounded output for the fill script to read
- Validation sheet: cross-form consistency checks
- Carryforwards sheet: items carrying to next year

Design principles:
- ALL computation happens in Python with exact precision (no rounding
  until the Form Values sheet)
- Each computed cell records its formula as a string for auditing
- The fill script reads from the Form Values sheet only
- A CPA can open the xlsx and trace any number back to its source

Usage:
    from build_workbook import TaxWorkbook

    wb = TaxWorkbook(tax_year=2025, taxpayer="Jane Doe")

    wb.source_data("W-2", [
        ("Box 1 - Wages", 92347.50, "w-2.pdf"),
        ("Box 2 - Federal withheld", 14283.00, "w-2.pdf"),
    ])

    wb.computation("Rental Property", [
        Row("Unit A sq ft fraction", val=800/2000, formula="800 / 2000"),
        Row("Unit A mortgage interest", val=24150 * 800/2000 * 0.5,
            formula="SourceData[Mortgage interest] * 800/2000 * 0.5"),
    ])

    wb.federal_return([
        Row("1a", "Wages", val=92348, formula="ROUND(SourceData[W-2 Box 1])"),
    ])

    wb.form_values("1040", [
        FormField("1a", "Wages", val=92348, pdf_field="f1_47"),
    ])

    wb.validate([
        Check("Sch D line 16 == 1040 line 7",
              expected=7834, actual=7834),
    ])

    wb.save("work/tax_computations.xlsx")
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from typing import Any, Optional


# ---------------------------------------------------------------------------
# Data classes for structured input
# ---------------------------------------------------------------------------

@dataclass
class Row:
    """A single row in a computation or return sheet."""
    label: str
    val: Any = None
    formula: str = ""
    notes: str = ""
    line: str = ""          # form line number (e.g., "1a", "7")
    is_subtotal: bool = False
    is_total: bool = False


@dataclass
class FormField:
    """A single field in the Form Values sheet."""
    line: str               # form line number
    description: str
    val: Any                # rounded value (what goes on the form)
    pdf_field: str          # PDF field name for fill script
    source: str = ""        # reference to computation cell


@dataclass
class Check:
    """A single validation check."""
    description: str
    expected: Any
    actual: Any
    tolerance: float = 0.5  # allow rounding differences

    @property
    def passed(self) -> bool:
        if self.expected is None or self.actual is None:
            return False
        try:
            return abs(float(self.expected) - float(self.actual)) <= self.tolerance
        except (ValueError, TypeError):
            return str(self.expected) == str(self.actual)


@dataclass
class Carryforward:
    """An item that carries forward to next year."""
    item: str
    amount: Any
    originated: str = ""
    expires: str = ""
    notes: str = ""


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
CURRENCY_FMT = '#,##0.00'
CURRENCY_ROUND_FMT = '#,##0'
PCT_FMT = '0.00000%'
THIN_BORDER = Border(
    bottom=Side(style='thin')
)


def _style_header_row(ws, ncols):
    """Apply header styling to row 1."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.freeze_panes = "A2"


def _auto_width(ws, min_width=10, max_width=50):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _format_value(cell, val):
    """Apply number formatting based on value type."""
    cell.value = val
    if isinstance(val, float):
        if abs(val) < 1 and val != 0:
            cell.number_format = PCT_FMT
        else:
            cell.number_format = CURRENCY_FMT
    elif isinstance(val, int):
        cell.number_format = CURRENCY_ROUND_FMT


# ---------------------------------------------------------------------------
# TaxWorkbook
# ---------------------------------------------------------------------------

def preflight_check(work_dir, forms_dir, filing_status="single"):
    """Verify required artifacts exist and extract tax tables from forms.

    Call this BEFORE building the computation workbook. It:
    1. Checks that instruction-notes files exist (proving instructions were read)
    2. Extracts tax table values from the actual downloaded forms
    3. Returns the extracted values for use in the build script

    Args:
        work_dir: Path to work/ directory
        forms_dir: Path to forms/ directory
        filing_status: "single", "mfj", "hoh"

    Returns:
        Dict of extracted tax table values from the forms themselves.

    Raises:
        FileNotFoundError: If required artifacts are missing
    """
    from extract_tax_tables import extract_all

    # Check for instruction-notes artifacts
    required_notes = [
        ("state_instructions_notes.txt", "State form instructions (Phase 1b)"),
    ]
    missing = []
    for filename, description in required_notes:
        path = os.path.join(work_dir, filename)
        if not os.path.exists(path):
            missing.append(f"  {filename} — {description}")
        elif os.path.getsize(path) < 100:
            missing.append(f"  {filename} — exists but too small ({os.path.getsize(path)} bytes). Did you actually read the instructions?")

    if missing:
        msg = "PREFLIGHT FAILED — required instruction-notes artifacts missing:\n"
        msg += "\n".join(missing)
        msg += "\n\nThe skill requires reading form instructions BEFORE computing."
        msg += "\nCreate these files by reading the instructions, then re-run."
        raise FileNotFoundError(msg)

    # Extract values from the actual forms
    extracted = extract_all(forms_dir)

    if not extracted:
        raise RuntimeError(f"No values extracted from forms in {forms_dir}. Are blank forms downloaded?")

    print(f"  Preflight: extracted values from {len(extracted)} forms")
    for section, values in extracted.items():
        count = len(values) - 1  # -1 for _source
        source = values.get("_source", "?")
        print(f"    {section}: {count} values from {source}")

    return extracted


class TaxWorkbook:
    """Builder for a tax computation workbook."""

    def __init__(self, tax_year: int, taxpayer: str):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # remove default sheet
        self.tax_year = tax_year
        self.taxpayer = taxpayer
        self._sheets_created = set()

    # -- Source Data --------------------------------------------------------

    def source_data(self, category: str, items: list[tuple]):
        """Add rows to the Source Data sheet.

        Args:
            category: Group name (e.g., "W-2", "Schwab 1099-B")
            items: List of (item_name, value, source_document) tuples
        """
        if "Source Data" not in self._sheets_created:
            ws = self.wb.create_sheet("Source Data", 0)
            ws.append(["Category", "Item", "Value", "Source"])
            _style_header_row(ws, 4)
            self._sheets_created.add("Source Data")
        else:
            ws = self.wb["Source Data"]

        for item_name, value, source in items:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=category)
            ws.cell(row=row_num, column=2, value=item_name)
            _format_value(ws.cell(row=row_num, column=3), value)
            ws.cell(row=row_num, column=4, value=source)

        _auto_width(ws)

    # -- Tax Tables --------------------------------------------------------

    def tax_tables(self, tables: dict[str, list[tuple]]):
        """Add a Tax Tables sheet.

        Args:
            tables: Dict of {section_name: [(item, value, source), ...]}
        """
        ws = self.wb.create_sheet("Tax Tables")
        ws.append(["Section", "Item", "Value", "Source"])
        _style_header_row(ws, 4)
        self._sheets_created.add("Tax Tables")

        for section, items in tables.items():
            for item_name, value, source in items:
                row_num = ws.max_row + 1
                ws.cell(row=row_num, column=1, value=section)
                ws.cell(row=row_num, column=2, value=item_name)
                _format_value(ws.cell(row=row_num, column=3), value)
                ws.cell(row=row_num, column=4, value=source)

        _auto_width(ws)

    # -- Computation Sheets ------------------------------------------------

    def computation(self, name: str, rows: list[Row]):
        """Add a computation sheet.

        Args:
            name: Sheet name (e.g., "Capital Gains", "Rental Property")
            rows: List of Row objects
        """
        ws = self.wb.create_sheet(name)
        ws.append(["Line", "Description", "Value", "Formula", "Notes"])
        _style_header_row(ws, 5)
        self._sheets_created.add(name)

        for row in rows:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=row.line)
            ws.cell(row=row_num, column=2, value=row.label)
            _format_value(ws.cell(row=row_num, column=3), row.val)
            ws.cell(row=row_num, column=4, value=row.formula)
            ws.cell(row=row_num, column=5, value=row.notes)

            if row.is_total:
                for col in range(1, 6):
                    c = ws.cell(row=row_num, column=col)
                    c.font = TOTAL_FONT
                    c.fill = TOTAL_FILL
                    c.border = THIN_BORDER
            elif row.is_subtotal:
                for col in range(1, 6):
                    c = ws.cell(row=row_num, column=col)
                    c.font = SUBTOTAL_FONT
                    c.fill = SUBTOTAL_FILL

        _auto_width(ws)

    # -- Federal / State Return Sheets -------------------------------------

    def federal_return(self, rows: list[Row]):
        """Add a Federal Return sheet (one row per 1040 line)."""
        self.computation(f"Federal Return ({self.tax_year})", rows)

    def state_return(self, state: str, rows: list[Row]):
        """Add a State Return sheet."""
        self.computation(f"{state} Return ({self.tax_year})", rows)

    # -- Form Values (output for fill script) ------------------------------

    def form_values(self, form_name: str, fields: list[FormField]):
        """Add rows to the Form Values sheet.

        Args:
            form_name: Form identifier (e.g., "1040", "Schedule D", "CA 540")
            fields: List of FormField objects
        """
        if "Form Values" not in self._sheets_created:
            ws = self.wb.create_sheet("Form Values")
            ws.append(["Form", "Line", "Description", "Value", "PDF Field", "Source"])
            _style_header_row(ws, 6)
            self._sheets_created.add("Form Values")
        else:
            ws = self.wb["Form Values"]

        for f in fields:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=form_name)
            ws.cell(row=row_num, column=2, value=f.line)
            ws.cell(row=row_num, column=3, value=f.description)
            _format_value(ws.cell(row=row_num, column=4), f.val)
            ws.cell(row=row_num, column=5, value=f.pdf_field)
            ws.cell(row=row_num, column=6, value=f.source)

        _auto_width(ws)

    # -- Validation --------------------------------------------------------

    def validate(self, checks: list[Check]):
        """Add a Validation sheet with cross-form checks."""
        ws = self.wb.create_sheet("Validation")
        ws.append(["Check", "Expected", "Actual", "Status"])
        _style_header_row(ws, 4)
        self._sheets_created.add("Validation")

        all_passed = True
        for check in checks:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=check.description)
            _format_value(ws.cell(row=row_num, column=2), check.expected)
            _format_value(ws.cell(row=row_num, column=3), check.actual)

            status = "PASS" if check.passed else "FAIL"
            status_cell = ws.cell(row=row_num, column=4, value=status)
            status_cell.fill = PASS_FILL if check.passed else FAIL_FILL
            status_cell.font = Font(bold=True)

            if not check.passed:
                all_passed = False

        _auto_width(ws)
        return all_passed

    # -- Carryforwards -----------------------------------------------------

    def carryforwards(self, items: list[Carryforward]):
        """Add a Carryforwards sheet."""
        ws = self.wb.create_sheet("Carryforwards")
        ws.append(["Item", "Amount", "Originated", "Expires", "Notes"])
        _style_header_row(ws, 5)
        self._sheets_created.add("Carryforwards")

        for item in items:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=item.item)
            _format_value(ws.cell(row=row_num, column=2), item.amount)
            ws.cell(row=row_num, column=3, value=item.originated)
            ws.cell(row=row_num, column=4, value=item.expires)
            ws.cell(row=row_num, column=5, value=item.notes)

        _auto_width(ws)

    # -- Save --------------------------------------------------------------

    def save(self, path: str):
        """Save the workbook to disk."""
        self.wb.save(path)
        print(f"  Saved: {path}")
        print(f"  Sheets: {', '.join(self.wb.sheetnames)}")


# ---------------------------------------------------------------------------
# Utility functions for tax computations
# ---------------------------------------------------------------------------

# MACRS depreciation rates (IRS Publication 946, Table A-6)
# 27.5-year residential rental property, mid-month convention
MACRS_27_5 = {
    # month_placed_in_service: {year: rate}
    1:  {1: 3.485, 2: 3.636},
    2:  {1: 3.182, 2: 3.636},
    3:  {1: 2.879, 2: 3.636},
    4:  {1: 2.576, 2: 3.636},
    5:  {1: 2.273, 2: 3.636},
    6:  {1: 1.970, 2: 3.636},
    7:  {1: 1.667, 2: 3.636},
    8:  {1: 1.364, 2: 3.636},
    9:  {1: 1.061, 2: 3.636},
    10: {1: 0.758, 2: 3.636},
    11: {1: 0.455, 2: 3.636},
    12: {1: 0.152, 2: 3.636},
}

# MACRS 5-year property, half-year convention (IRS Pub 946, Table A-1)
MACRS_5 = {1: 20.00, 2: 32.00, 3: 19.20, 4: 11.52, 5: 11.52, 6: 5.76}

# MACRS 7-year property, half-year convention
MACRS_7 = {1: 14.29, 2: 24.49, 3: 17.49, 4: 12.49, 5: 8.93, 6: 8.92, 7: 8.93, 8: 4.46}


def macrs_rate(recovery_years: int, year: int, month_placed: int = None) -> float:
    """Look up MACRS depreciation rate as a decimal (e.g., 0.03636).

    Args:
        recovery_years: 5, 7, or 27.5
        year: Depreciation year (1, 2, 3, ...)
        month_placed: Month placed in service (1-12), required for 27.5-year

    Returns:
        Depreciation rate as a decimal (e.g., 0.03636 for 3.636%)
    """
    if recovery_years == 27.5:
        if month_placed is None:
            raise ValueError("month_placed required for 27.5-year property")
        pct = MACRS_27_5[month_placed].get(year, 3.636)
        return pct / 100
    elif recovery_years == 5:
        return MACRS_5.get(year, 0) / 100
    elif recovery_years == 7:
        return MACRS_7.get(year, 0) / 100
    else:
        raise ValueError(f"Unsupported recovery period: {recovery_years}")


def compute_tax(taxable_income: float, brackets: list[tuple]) -> float:
    """Compute tax using progressive brackets.

    Args:
        taxable_income: Taxable income amount
        brackets: List of (threshold, rate) tuples, sorted ascending.
                  e.g., [(0, 0.10), (11925, 0.12), (48475, 0.22), ...]

    Returns:
        Total tax computed across all brackets.
    """
    tax = 0.0
    prev_threshold = 0
    for i, (threshold, rate) in enumerate(brackets):
        if i == 0:
            prev_threshold = threshold
            continue
        if taxable_income <= prev_threshold:
            break
        taxable_in_bracket = min(taxable_income, threshold) - prev_threshold
        if taxable_in_bracket > 0:
            tax += taxable_in_bracket * brackets[i - 1][1]
        prev_threshold = threshold

    # Top bracket (no upper limit)
    if taxable_income > brackets[-1][0]:
        tax += (taxable_income - brackets[-1][0]) * brackets[-1][1]
    elif len(brackets) > 1 and taxable_income > brackets[-2][0]:
        # Handle the last bounded bracket
        pass  # already handled in loop

    return tax


def compute_tax_simple(taxable_income: float, brackets: list[tuple]) -> float:
    """Compute tax using brackets specified as (upper_bound, rate) pairs.

    The last bracket should use float('inf') as upper bound.

    Args:
        taxable_income: Taxable income
        brackets: [(upper_bound, rate), ...] e.g.,
                  [(11925, 0.10), (48475, 0.12), ..., (float('inf'), 0.37)]

    Returns:
        Total tax
    """
    tax = 0.0
    prev = 0
    for upper, rate in brackets:
        if taxable_income <= prev:
            break
        taxable_in_bracket = min(taxable_income, upper) - prev
        tax += taxable_in_bracket * rate
        prev = upper
    return tax


def round_dollar(val: float) -> int:
    """Round to nearest whole dollar (IRS standard rounding)."""
    return round(val)
